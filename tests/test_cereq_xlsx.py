"""Tests pour le parser xlsx Céreq (D11 Axe 1).

Focus : mapping code → niveau OrientIA, jointure 3ans+6ans sur Code,
extraction métriques, gestion fichiers manquants.
"""
from __future__ import annotations

from pathlib import Path

import pytest
import openpyxl

from src.collect.cereq import (
    CEREQ_CODE_TO_NIVEAU,
    CEREQ_DOMAINE_HINTS,
    CereqDataMissing,
    _metrics_from_row,
    parse_cereq_xlsx,
)


REPO_ROOT = Path(__file__).resolve().parent.parent
REAL_XLSX = REPO_ROOT / "data" / "raw" / "OpenData_Cereq-Enq_Generation-Donnees_DIPLOME.xlsx"


def _make_fake_cereq_xlsx(tmp_path: Path) -> Path:
    """Mini xlsx Céreq-like pour tests unitaires (pas de dépendance I/O réelle)."""
    wb = openpyxl.Workbook()
    # Sheet 1 : Info (doit être ignorée)
    ws_info = wb.active
    ws_info.title = "Information_3ans"
    ws_info.append(["Présentation", None])
    # Sheet 2 : Données 3ans
    ws3 = wb.create_sheet("Données_3ans")
    ws3.append([
        "Code", "Libelle_Menu", "Libelle complet",
        "eff_pond_g1", "taux_emploi", "revenu_travail",
    ])
    ws3.append([310, "CAP et autres", "CAP et autres diplômes de niveau 3", 43127, 61, 1330])
    ws3.append([511, "Ensemble", "Masters", 93448, 85, 2080])
    ws3.append([532, "Domaine Santé", "Doctorats du domaine de la Santé", 5000, 95, 4500])
    # Sheet 3 : Données 6ans (join sur Code)
    ws6 = wb.create_sheet("Données_6ans")
    ws6.append([
        "Code", "Libelle_Menu", "Libelle complet",
        "taux_emploi", "revenu_travail",
    ])
    ws6.append([310, "CAP et autres", "CAP et autres diplômes de niveau 3", 69, 1500])
    ws6.append([511, "Ensemble", "Masters", 91, 2500])
    # 532 absent du 6ans pour tester la robustesse
    path = tmp_path / "cereq_mini.xlsx"
    wb.save(path)
    return path


def test_parse_cereq_xlsx_mini_fixture(tmp_path):
    path = _make_fake_cereq_xlsx(tmp_path)
    entries = parse_cereq_xlsx(path)
    assert len(entries) == 3
    codes = {e["code"] for e in entries}
    assert codes == {310, 511, 532}


def test_niveau_mapping_from_code(tmp_path):
    path = _make_fake_cereq_xlsx(tmp_path)
    entries = {e["code"]: e for e in parse_cereq_xlsx(path)}
    assert entries[310]["niveau"] == "cap-bep"
    assert entries[511]["niveau"] == "bac+5"
    assert entries[532]["niveau"] == "bac+8"


def test_domaine_hints_apply(tmp_path):
    path = _make_fake_cereq_xlsx(tmp_path)
    entries = {e["code"]: e for e in parse_cereq_xlsx(path)}
    # 310 : pas de domaine (générique CAP)
    assert entries[310]["domaine"] is None
    # 511 : Masters Ensemble (pas de domaine — générique)
    assert entries[511]["domaine"] is None
    # 532 : Doctorats Santé → domaine sante
    assert entries[532]["domaine"] == "sante"


def test_horizon_3ans_6ans_joined_when_both_present(tmp_path):
    path = _make_fake_cereq_xlsx(tmp_path)
    entries = {e["code"]: e for e in parse_cereq_xlsx(path)}
    assert entries[511]["horizon_3ans"]["taux_emploi"] == 85
    assert entries[511]["horizon_6ans"]["taux_emploi"] == 91


def test_horizon_6ans_none_when_code_absent_from_6ans(tmp_path):
    path = _make_fake_cereq_xlsx(tmp_path)
    entries = {e["code"]: e for e in parse_cereq_xlsx(path)}
    assert entries[532]["horizon_3ans"]["taux_emploi"] == 95
    assert entries[532]["horizon_6ans"] is None


def test_metrics_from_row_drops_administrative_columns():
    row = {
        "Code": 511, "Libelle_Menu": "X", "Libelle complet": "Y",
        "taux_emploi": 85, "revenu_travail": 2080, "eff_pond_g1": 93448,
    }
    metrics = _metrics_from_row(row)
    assert "Code" not in metrics
    assert "Libelle_Menu" not in metrics
    assert "Libelle complet" not in metrics
    assert metrics == {"taux_emploi": 85, "revenu_travail": 2080, "eff_pond_g1": 93448}


def test_metrics_from_row_drops_none_values():
    row = {"taux_emploi": 85, "revenu_travail": None, "taux_chomage": 10}
    metrics = _metrics_from_row(row)
    assert "revenu_travail" not in metrics
    assert metrics == {"taux_emploi": 85, "taux_chomage": 10}


def test_missing_file_raises_cereq_data_missing(tmp_path):
    missing = tmp_path / "does_not_exist.xlsx"
    with pytest.raises(CereqDataMissing):
        parse_cereq_xlsx(missing)


def test_cereq_code_table_complete():
    # Sanity check : hiérarchie Céreq 300-533 doit être couverte
    expected_codes = [
        310, 311, 312, 313, 320, 321, 322, 323, 330, 331, 332, 333, 340,
        410, 411, 412, 413, 420, 430, 431, 432, 433, 440, 441, 442, 443,
        510, 511, 512, 513, 520, 521, 522, 523, 530, 531, 532, 533,
    ]
    for c in expected_codes:
        assert c in CEREQ_CODE_TO_NIVEAU, f"Code Céreq {c} manquant dans mapping"


@pytest.mark.skipif(not REAL_XLSX.exists(), reason="XLSX Céreq réel absent")
def test_parse_real_xlsx_produces_expected_entries():
    """Intégration : parse le xlsx Céreq réel livré par Matteo."""
    entries = parse_cereq_xlsx(REAL_XLSX)
    assert 40 <= len(entries) <= 50, f"Attendu ~43 entrées, obtenu {len(entries)}"
    # Au moins 1 Master
    assert any(e["code"] == 511 and e["niveau"] == "bac+5" for e in entries)
    # Au moins 1 CAP
    assert any(e["code"] == 310 and e["niveau"] == "cap-bep" for e in entries)
    # Les 2 horizons doivent être présents pour les codes principaux
    master = next(e for e in entries if e["code"] == 511)
    assert master["horizon_3ans"] is not None
    assert master["horizon_6ans"] is not None
    assert "taux_emploi" in master["horizon_3ans"]
