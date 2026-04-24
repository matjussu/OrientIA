"""Tests pour `src/collect/dares_metiers_2030.py` — scaffold prospectif."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.collect.dares_metiers_2030 import (
    DaresDataMissing,
    DaresLayoutUnexpected,
    FAP_PREFIX_TO_DOMAINE,
    ParserConfig,
    _safe_int,
    collect_dares_metiers_2030,
    fap_to_domaine,
    parse_dares_fap_xlsx,
    parse_dares_region_xlsx,
)


# --- Mappings ---


def test_fap_prefix_to_domaine_tous_secteurs_couverts():
    # 21 secteurs attendus (A, B, C, D, E, F, G, H, J, K, L, M, N, P, Q, R, S, T, U, V, W)
    assert len(FAP_PREFIX_TO_DOMAINE) >= 20


@pytest.mark.parametrize(
    "fap,expected",
    [
        ("E1Z80", "data_ia"),
        ("E1", "data_ia"),
        ("B0Z40", "btp"),
        ("V1Z20", "sante"),
        ("M0Z00", "eco_gestion"),
        ("Z99", None),  # lettre inconnue
        (None, None),
        ("", None),
    ],
)
def test_fap_to_domaine(fap, expected):
    assert fap_to_domaine(fap) == expected


def test_safe_int_fr_formatting():
    assert _safe_int("38\u202f000") == 38000
    assert _safe_int("1 234,5") == 1234
    assert _safe_int("") is None
    assert _safe_int(None) is None


# --- FAP parser fixture ---


def _make_fap_fixture(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projections FAP"
    ws.append([
        "Code FAP", "Libellé FAP", "Création postes 2020-2030",
        "Départs retraite 2020-2030", "Total à pourvoir", "Scénario",
    ])
    ws.append(["E1Z80", "Ingénieurs en informatique", 50000, 30000, 80000, "central"])
    ws.append(["B0Z40", "Ouvriers du BTP", 20000, 80000, 100000, "central"])
    ws.append(["V1Z20", "Médecins", 15000, 25000, 40000, "central"])
    path = tmp_path / "Metiers2030_FAP87_Projections.xlsx"
    wb.save(path)
    return path


def test_parse_dares_fap_xlsx_basic(tmp_path):
    path = _make_fap_fixture(tmp_path)
    entries = parse_dares_fap_xlsx(path)
    assert len(entries) == 3
    first = entries[0]
    assert first["source"] == "dares_metiers_2030"
    assert first["horizon"] == "2020-2030"
    assert first["fap_code"] == "E1Z80"
    assert first["fap_libelle"] == "Ingénieurs en informatique"
    assert first["domaine_orientia"] == "data_ia"
    assert first["postes_creation"] == 50000
    assert first["postes_depart_retraite"] == 30000
    assert first["postes_a_pourvoir_total"] == 80000
    assert first["scenario"] == "central"


def test_parse_dares_fap_domaine_mapping(tmp_path):
    path = _make_fap_fixture(tmp_path)
    entries = {e["fap_code"]: e for e in parse_dares_fap_xlsx(path)}
    assert entries["E1Z80"]["domaine_orientia"] == "data_ia"
    assert entries["B0Z40"]["domaine_orientia"] == "btp"
    assert entries["V1Z20"]["domaine_orientia"] == "sante"


def test_parse_dares_fap_missing_file_raises(tmp_path):
    with pytest.raises(DaresDataMissing):
        parse_dares_fap_xlsx(tmp_path / "nope.xlsx")


def test_parse_dares_fap_missing_columns_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuille"
    ws.append(["Autre", "Sans rapport"])
    ws.append(["x", "y"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(DaresLayoutUnexpected) as exc:
        parse_dares_fap_xlsx(path)
    assert "fap_code" in str(exc.value) or "libelle" in str(exc.value)


# --- Region parser fixture ---


def _make_region_fixture(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Régions"
    ws.append([
        "Région", "Secteur",
        "Création postes 2020-2030", "Départs retraite 2020-2030",
        "Total à pourvoir",
    ])
    ws.append(["Île-de-France", "Ensemble", 500000, 800000, 1300000])
    ws.append(["Nouvelle-Aquitaine", "Ensemble", 150000, 300000, 450000])
    path = tmp_path / "Metiers2030_Region.xlsx"
    wb.save(path)
    return path


def test_parse_dares_region_basic(tmp_path):
    path = _make_region_fixture(tmp_path)
    entries = parse_dares_region_xlsx(path)
    assert len(entries) == 2
    first = entries[0]
    assert first["source"] == "dares_metiers_2030_region"
    assert first["region"] == "Île-de-France"
    assert first["postes_a_pourvoir_total"] == 1300000


def test_parse_dares_region_missing_columns(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.append(["Colonne autre", "Foo"])
    wb.active.append(["x", "y"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(DaresLayoutUnexpected):
        parse_dares_region_xlsx(path)


# --- collect_dares_metiers_2030 (entrée principale) ---


def test_collect_dares_missing_raw_dir(tmp_path):
    missing_dir = tmp_path / "doesnt_exist"
    with pytest.raises(DaresDataMissing):
        collect_dares_metiers_2030(raw_dir=missing_dir)


def test_collect_dares_empty_dir(tmp_path):
    raw_dir = tmp_path / "dares"
    raw_dir.mkdir()
    with pytest.raises(DaresDataMissing):
        collect_dares_metiers_2030(raw_dir=raw_dir, save=False)


def test_collect_dares_picks_up_fap_and_region(tmp_path):
    raw_dir = tmp_path / "dares"
    raw_dir.mkdir()
    _make_fap_fixture(raw_dir)
    _make_region_fixture(raw_dir)
    out = collect_dares_metiers_2030(raw_dir=raw_dir, save=False)
    # 3 FAP + 2 région
    assert len(out) == 5
    sources = {e["source"] for e in out}
    assert sources == {"dares_metiers_2030", "dares_metiers_2030_region"}


def test_parser_config_custom_sheet(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Info"
    wb.active.append(["Skip"])
    ws2 = wb.create_sheet("MesDonnees")
    ws2.append(["Code FAP", "Libellé FAP", "Total à pourvoir"])
    ws2.append(["E1", "Info", 100000])
    path = tmp_path / "custom.xlsx"
    wb.save(path)
    out = parse_dares_fap_xlsx(path, config=ParserConfig(sheet_name="MesDonnees"))
    assert len(out) == 1
    assert out[0]["fap_code"] == "E1"
