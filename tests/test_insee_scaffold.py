"""Tests pour `src/collect/insee.py` scaffold D14 (ADR-039 phase c).

Focus : mappings PCS + diplôme → OrientIA, parsers xlsx avec fixtures
mini-reproductibles, gestion fichier/colonnes manquant(e)s.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.collect.insee import (
    DIPLOME_TO_NIVEAU,
    InseeDataMissing,
    InseeLayoutUnexpected,
    PCS_1CHIFFRE_TO_DOMAINE,
    PCS_2CHIFFRES_TO_DOMAINE,
    ParserConfig,
    _safe_int,
    _safe_ratio,
    collect_insee_stats,
    diplome_to_niveau,
    parse_enquete_emploi_xlsx,
    parse_salcs_xlsx,
    pcs_to_domaine,
)


# --- Mappings ---


def test_pcs_2_chiffres_priority():
    assert pcs_to_domaine("38") == "cadres_ingenieurs"
    assert pcs_to_domaine("384000") == "cadres_ingenieurs"  # granularité préfixe 2 chiffres


def test_pcs_1_chiffre_fallback():
    # Code 2-chiffres "39" pas dans le mapping granulaire → fallback sur "3"
    assert pcs_to_domaine("39") == "cadres_prof_sup"
    assert pcs_to_domaine("5") == "employes"


def test_pcs_unknown_returns_none():
    assert pcs_to_domaine(None) is None
    assert pcs_to_domaine("") is None
    assert pcs_to_domaine("999") is None


def test_diplome_to_niveau_common_labels():
    assert diplome_to_niveau("Bac+5 ou plus") == "bac+5"
    assert diplome_to_niveau("Bac + 5") == "bac+5"
    assert diplome_to_niveau("Master professionnel") == "bac+5"
    assert diplome_to_niveau("Licence") == "bac+3"
    assert diplome_to_niveau("CAP - BEP") == "cap-bep"
    assert diplome_to_niveau("Baccalauréat général") == "bac"
    assert diplome_to_niveau("Doctorat") == "bac+8"


def test_diplome_to_niveau_priority_longer_match():
    # "Bac+5" doit l'emporter sur "Baccalauréat" générique
    assert diplome_to_niveau("Bac+5 ou plus (master)") == "bac+5"


def test_diplome_to_niveau_unknown_returns_none():
    assert diplome_to_niveau(None) is None
    assert diplome_to_niveau("") is None
    assert diplome_to_niveau("Label exotique inconnu") is None


# --- Safe parsers ---


def test_safe_int_handles_insee_formatting():
    # INSEE utilise parfois espace insécable U+202F comme séparateur milliers
    assert _safe_int("38\u202f000") == 38000
    assert _safe_int("1 850 €") == 1850
    assert _safe_int("1850,50") == 1850
    assert _safe_int("") is None
    assert _safe_int(None) is None


def test_safe_ratio_percentage_vs_float():
    assert _safe_ratio("82,5%") == pytest.approx(0.825)
    assert _safe_ratio("82.5") == pytest.approx(0.825)
    assert _safe_ratio(0.825) == 0.825
    assert _safe_ratio(82.5) == pytest.approx(0.825)  # >1.5 → pct
    assert _safe_ratio(0.5) == 0.5  # <1.5 → déjà ratio
    assert _safe_ratio(None) is None


# --- SALCS xlsx parser ---


def _make_salcs_fixture(tmp_path: Path) -> Path:
    """Mini xlsx reproduisant un layout INSEE SALCS typique."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Figure 1"
    ws.append([
        "PCS", "Libellé PCS", "Tranche d'âge", "Sexe",
        "Salaire net EQTP médian (€/an)", "1er décile", "9e décile", "Effectif",
    ])
    ws.append(["38", "Ingénieurs et cadres techniques", "25-29 ans", "Tous", 38000, 28000, 55000, 124000])
    ws.append(["38", "Ingénieurs et cadres techniques", "30-34 ans", "Tous", 45000, 32000, 65000, 198000])
    ws.append(["47", "Techniciens", "25-29 ans", "Tous", 28000, 22000, 38000, 85000])
    ws.append(["54", "Employés administratifs", "25-29 ans", "Tous", 22000, 18000, 28000, 162000])
    path = tmp_path / "insee_salaires_pcs_age_sexe_2022.xlsx"
    wb.save(path)
    return path


def test_parse_salcs_xlsx_basic(tmp_path):
    path = _make_salcs_fixture(tmp_path)
    entries = parse_salcs_xlsx(path)
    assert len(entries) == 4
    # Spot check premier record
    first = entries[0]
    assert first["source"] == "insee_salcs"
    assert first["pcs"] == "38"
    assert first["pcs_label"] == "Ingénieurs et cadres techniques"
    assert first["domaine_orientia"] == "cadres_ingenieurs"
    assert first["tranche_age"] == "25-29 ans"
    assert first["salaire_eqtp_net_median_annuel"] == 38000
    assert first["salaire_eqtp_net_d1"] == 28000
    assert first["salaire_eqtp_net_d9"] == 55000
    assert first["effectif"] == 124000


def test_parse_salcs_millesime_inferred_from_filename(tmp_path):
    path = _make_salcs_fixture(tmp_path)
    entries = parse_salcs_xlsx(path)
    assert entries[0]["millesime"] == "2022"


def test_parse_salcs_missing_file_raises(tmp_path):
    with pytest.raises(InseeDataMissing):
        parse_salcs_xlsx(tmp_path / "nope.xlsx")


def test_parse_salcs_missing_required_columns_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Figure 1"
    ws.append(["Autre colonne", "Foo"])
    ws.append(["x", "y"])
    path = tmp_path / "insee_bad.xlsx"
    wb.save(path)
    with pytest.raises(InseeLayoutUnexpected) as exc:
        parse_salcs_xlsx(path)
    assert "pcs" in str(exc.value) or "salaire_median" in str(exc.value)


def test_parse_salcs_custom_sheet_config(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Info"
    wb.active.append(["Info sheet", "to be skipped"])
    ws2 = wb.create_sheet("Tableau1")
    ws2.append(["PCS", "Salaire médian"])
    ws2.append(["31", 50000])
    path = tmp_path / "insee_custom.xlsx"
    wb.save(path)
    entries = parse_salcs_xlsx(path, config=ParserConfig(sheet_name="Tableau1"))
    assert len(entries) == 1
    assert entries[0]["pcs"] == "31"
    assert entries[0]["salaire_eqtp_net_median_annuel"] == 50000


# --- Enquête Emploi xlsx parser ---


def _make_emploi_fixture(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taux d'emploi"
    ws.append([
        "Diplôme", "Tranche d'âge", "Sexe",
        "Taux d'emploi (%)", "Taux de chômage (%)", "Taux d'activité (%)",
    ])
    ws.append(["Bac+5 ou plus", "25-34", "Tous", 88.5, 6.2, 94.3])
    ws.append(["Bac+3 ou Bac+4", "25-34", "Tous", 82.1, 8.5, 89.7])
    ws.append(["Bac+2", "25-34", "Tous", 78.0, 10.2, 86.8])
    ws.append(["Baccalauréat", "25-34", "Tous", 68.3, 15.5, 80.8])
    ws.append(["CAP - BEP", "25-34", "Tous", 62.0, 18.0, 75.6])
    path = tmp_path / "insee_taux_emploi_diplome_2024T4.xlsx"
    wb.save(path)
    return path


def test_parse_enquete_emploi_basic(tmp_path):
    path = _make_emploi_fixture(tmp_path)
    entries = parse_enquete_emploi_xlsx(path)
    assert len(entries) == 5
    first = entries[0]
    assert first["source"] == "insee_enquete_emploi"
    assert first["diplome_label"] == "Bac+5 ou plus"
    assert first["niveau_orientia"] == "bac+5"
    assert first["tranche_age"] == "25-34"
    assert first["taux_emploi"] == pytest.approx(0.885)
    assert first["taux_chomage"] == pytest.approx(0.062)
    assert first["taux_activite"] == pytest.approx(0.943)


def test_parse_enquete_emploi_niveau_mapping(tmp_path):
    path = _make_emploi_fixture(tmp_path)
    entries = parse_enquete_emploi_xlsx(path)
    niveaux = [e["niveau_orientia"] for e in entries]
    assert niveaux == ["bac+5", "bac+3", "bac+2", "bac", "cap-bep"]


def test_parse_enquete_emploi_periode_inferred(tmp_path):
    path = _make_emploi_fixture(tmp_path)
    entries = parse_enquete_emploi_xlsx(path)
    # Pattern "2024T4" détecté → périodes trimestrielles
    assert entries[0]["periode"] == "2024T4"


def test_parse_enquete_emploi_missing_file_raises(tmp_path):
    with pytest.raises(InseeDataMissing):
        parse_enquete_emploi_xlsx(tmp_path / "nope.xlsx")


def test_parse_enquete_emploi_missing_required_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Other", "Unrelated"])
    ws.append(["x", "y"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(InseeLayoutUnexpected):
        parse_enquete_emploi_xlsx(path)


# --- collect_insee_stats (entrée principale) ---


def test_collect_insee_stats_missing_raw_dir(tmp_path):
    raw_dir = tmp_path / "doesnt_exist"
    with pytest.raises(InseeDataMissing) as exc:
        collect_insee_stats(raw_dir=raw_dir)
    assert "absent" in str(exc.value).lower()


def test_collect_insee_stats_picks_up_both_datasets(tmp_path):
    raw_dir = tmp_path / "insee"
    raw_dir.mkdir()
    # Crée les 2 fixtures dans le répertoire raw
    _make_salcs_fixture(raw_dir).rename(raw_dir / "insee_salaires_pcs_2022.xlsx")
    _make_emploi_fixture(raw_dir).rename(raw_dir / "insee_taux_emploi_2024.xlsx")
    result = collect_insee_stats(raw_dir=raw_dir, save=False)
    assert len(result["salcs"]) == 4
    assert len(result["enquete_emploi"]) == 5


def test_collect_insee_stats_empty_dir(tmp_path):
    raw_dir = tmp_path / "insee"
    raw_dir.mkdir()
    with pytest.raises(InseeDataMissing) as exc:
        collect_insee_stats(raw_dir=raw_dir, save=False)
    assert "aucun xlsx" in str(exc.value).lower()


def test_mappings_tables_non_empty():
    """Sanity check : les tables de mapping ont des entrées."""
    assert len(PCS_1CHIFFRE_TO_DOMAINE) == 6
    assert len(PCS_2CHIFFRES_TO_DOMAINE) > 20
    assert len(DIPLOME_TO_NIVEAU) > 10
