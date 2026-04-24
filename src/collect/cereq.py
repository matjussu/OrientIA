"""Céreq Enquêtes Génération — insertion professionnelle jeunes diplômés (D11 Axe 1).

Source : https://www.cereq.fr/datavisualisation/insertion-professionnelle-des-jeunes/
Licence : statistique publique / Etalab 2.0 (données agrégées publiques).
Scope élargi ADR-039 : phase (c) insertion pro 17-25 ans.

**Format de livraison Matteo (2026-04-24)** : Céreq expose les données
d'Enquête Génération via xlsx OpenData téléchargeables. Matteo commit
les fichiers directement dans `data/raw/` (ex :
`OpenData_Cereq-Enq_Generation-Donnees_DIPLOME.xlsx` — Enquête
Génération 2017, horizons 3 ans et 6 ans par code diplôme).

Pipeline :
1. Matteo commit le(s) xlsx dans `data/raw/` (via commit GitHub).
2. `parse_cereq_xlsx()` lit les 2 feuilles de données (3ans + 6ans),
   joint sur le `Code` Céreq, mappe vers niveau OrientIA.
3. Produit `data/processed/cereq_insertion_stats.json` : table de
   référence stats insertion par niveau + domaine (cross-linkable avec
   les fiches OrientIA via le champ `niveau`).

Alternative future : Enquête `Génération 2021` dès publication
(parser identique attendu — mêmes codes Céreq).

Le parser CSV (`parse_chiffres_cles_csv`) est conservé pour un futur
format publication Céreq (API v2, si déployée).
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Optional


RAW_DIR_CSV = Path("data/raw/cereq")  # legacy dir for CSV exports
RAW_DIR = Path("data/raw")  # xlsx OpenData Céreq (commité par Matteo)
PROCESSED_PATH = Path("data/processed/cereq_insertion_stats.json")

# Mapping hiérarchie code Céreq → niveau OrientIA.
# Codes Céreq (Enquête Génération) : hiérarchie 3 chiffres.
#   1xx  Ensemble
#   2xx  Non-diplômés
#   3xx  Secondaire (310 CAP, 320 Bac pro, 330 Bac techno, 340 Bac gen)
#   4xx  Supérieur court (410 BTS-DUT, 420 Santé/Social, 430 Lic pro,
#                         440 Bac+3/+4)
#   5xx  Supérieur long (510 Masters, 520 Écoles, 530 Doctorat)
CEREQ_CODE_TO_NIVEAU: dict[int, str] = {
    # Secondaire
    310: "cap-bep", 311: "cap-bep", 312: "cap-bep", 313: "cap-bep",
    320: "bac", 321: "bac", 322: "bac", 323: "bac",
    330: "bac", 331: "bac", 332: "bac", 333: "bac",
    340: "bac",
    # Supérieur court (bac+2 / bac+3)
    410: "bac+2", 411: "bac+2", 412: "bac+2", 413: "bac+2",
    420: "bac+2",  # Diplômes Bac+2 à Bac+4 Santé/Social → bac+2 par défaut
    430: "bac+3", 431: "bac+3", 432: "bac+3", 433: "bac+3",
    440: "bac+3", 441: "bac+3", 442: "bac+3", 443: "bac+3",
    # Supérieur long
    510: "bac+5", 511: "bac+5", 512: "bac+5", 513: "bac+5",
    520: "bac+5", 521: "bac+5", 522: "bac+5", 523: "bac+5",
    530: "bac+8", 531: "bac+8", 532: "bac+8", 533: "bac+8",
}

# Domaine OrientIA inféré depuis le libellé Céreq (spécialité industrielle /
# tertiaire / LSH / scientifique / Santé).
CEREQ_DOMAINE_HINTS: dict[int, str] = {
    312: "industriel", 322: "industriel", 332: "industriel", 412: "industriel",
    313: "tertiaire", 323: "tertiaire", 333: "tertiaire", 413: "tertiaire",
    432: "lsh_gestion_droit", 442: "lsh_gestion_droit", 512: "lsh_gestion_droit",
    433: "scientifique", 443: "scientifique", 513: "scientifique",
    522: "ecole_commerce", 523: "ecole_ingenieur",
    532: "sante", 533: "hors_sante",
    420: "sante_social",
}


class CereqDataMissing(Exception):
    """Les CSVs Céreq attendus ne sont pas présents dans `data/raw/cereq/`.

    Pointe vers `docs/TODO_MATTEO_APIS.md` pour la procédure de download
    manuel depuis le portail Céreq.
    """


# --- Schéma attendu après normalisation ---
#
# Chaque entrée = 1 combinaison (niveau, domaine, cohorte) avec :
# - niveau : "bac+2" / "bac+3" / "bac+5" / ...  (aligné OrientIA)
# - domaine : label Céreq traduit en domaine OrientIA
# - cohorte : "Generation 2017" / "Generation 2021" / ...
# - taux_emploi_3ans, taux_emploi_6ans : float 0-1
# - taux_cdi : float 0-1
# - salaire_median_embauche : int (EUR mensuel brut)
# - delai_premier_emploi_mois : int (médiane)
# - source_url : str (lien Céreq)


def _infer_niveau(cereq_label: str) -> Optional[str]:
    """Mapping label Céreq → niveau OrientIA.

    Céreq groupe typiquement par : "Bac+5", "Bac+3", "Bac+2 et moins", etc.
    """
    if not cereq_label:
        return None
    l = cereq_label.lower().strip()
    if "bac+5" in l or "bac +5" in l or "master" in l:
        return "bac+5"
    if "bac+3" in l or "bac +3" in l or "licence" in l:
        return "bac+3"
    if "bac+2" in l or "bac +2" in l or "bts" in l or "but" in l:
        return "bac+2"
    if "cap" in l or "bep" in l:
        return "cap-bep"
    if "bac" in l and "+" not in l:
        return "bac"
    return None


def parse_chiffres_cles_csv(csv_path: Path) -> list[dict[str, Any]]:
    """Parse un CSV Céreq type 'Chiffres clés par diplôme'.

    Le schéma exact dépend du CSV téléchargé. Cette fonction est
    volontairement **permissive** : elle mappe ce qu'elle peut
    reconnaître et retourne les champs inconnus tels quels sous `_extra`.
    """
    if not csv_path.exists():
        raise CereqDataMissing(
            f"CSV Céreq absent : {csv_path}. "
            "Cf docs/TODO_MATTEO_APIS.md §4 pour la procédure download manuel."
        )
    with open(csv_path, encoding="utf-8") as f:
        # Céreq publie parfois en ; (FR), parfois en , (EN). On sniff.
        sample = f.read(2048)
        f.seek(0)
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)

    out: list[dict[str, Any]] = []
    for row in rows:
        niveau_label = row.get("niveau_diplome") or row.get("Niveau") or row.get("Diplôme")
        domaine = row.get("domaine") or row.get("Domaine") or row.get("Spécialité")
        entry = {
            "source": "cereq",
            "cohorte": row.get("cohorte") or row.get("Génération") or "Generation 2017",
            "niveau": _infer_niveau(niveau_label or ""),
            "niveau_label": niveau_label,
            "domaine": domaine,
            "taux_emploi_3ans": _safe_float(row.get("taux_emploi_3_ans") or row.get("Taux emploi 3 ans")),
            "taux_emploi_6ans": _safe_float(row.get("taux_emploi_6_ans") or row.get("Taux emploi 6 ans")),
            "taux_cdi": _safe_float(row.get("taux_cdi") or row.get("% CDI")),
            "salaire_median_embauche": _safe_int(row.get("salaire_median") or row.get("Salaire médian")),
            "delai_premier_emploi_mois": _safe_int(row.get("delai_emploi_mois") or row.get("Délai accès premier emploi")),
            "_extra": {k: v for k, v in row.items() if k not in {
                "niveau_diplome", "Niveau", "Diplôme", "domaine", "Domaine",
                "Spécialité", "cohorte", "Génération", "taux_emploi_3_ans",
                "Taux emploi 3 ans", "taux_emploi_6_ans", "Taux emploi 6 ans",
                "taux_cdi", "% CDI", "salaire_median", "Salaire médian",
                "delai_emploi_mois", "Délai accès premier emploi",
            }},
        }
        out.append(entry)
    return out


def _safe_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val).replace(",", ".").replace("€", "").strip()))
    except (ValueError, TypeError):
        return None


def save_processed(
    entries: list[dict[str, Any]], path: Path = PROCESSED_PATH
) -> Path:
    """Dump les stats insertion normalisées (commité)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return path


def parse_cereq_xlsx(
    xlsx_path: Path, cohorte: str = "Generation 2017"
) -> list[dict[str, Any]]:
    """Parse un xlsx OpenData Céreq (Enquête Génération).

    Format attendu (sheets) :
        Information_3ans, Information_6ans : métadonnées enquête (ignorées)
        Données_3ans : Code, Libelle_Menu, Libelle complet, eff_pond_*,
                       taux_emploi, taux_chomage, revenu_travail, ...
        Données_6ans : idem à horizon 6 ans

    La jointure 3ans ↔ 6ans se fait sur le `Code` Céreq. Une entrée de
    sortie par code, contenant les 2 horizons.

    Retourne une liste d'entrées au schéma :
        {
          "source": "cereq",
          "cohorte": str,
          "code": int,
          "libelle_menu": str,
          "libelle_complet": str,
          "niveau": str | None,         # OrientIA niveau (via CEREQ_CODE_TO_NIVEAU)
          "domaine": str | None,         # OrientIA domaine (via CEREQ_DOMAINE_HINTS)
          "horizon_3ans": dict | None,   # stats à 3 ans (ou None si absent)
          "horizon_6ans": dict | None,   # stats à 6 ans
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CereqDataMissing(
            "Package `openpyxl` absent. Installer via `pip install openpyxl` "
            "ou s'assurer qu'il est dans requirements.txt."
        ) from exc

    if not xlsx_path.exists():
        raise CereqDataMissing(
            f"XLSX Céreq absent : {xlsx_path}. "
            "Cf docs/TODO_MATTEO_APIS.md §4 pour la procédure download."
        )

    wb = load_workbook(xlsx_path, data_only=True)
    sheets_3ans = next(
        (s for s in wb.sheetnames if "Données" in s and "3" in s), None
    )
    sheets_6ans = next(
        (s for s in wb.sheetnames if "Données" in s and "6" in s), None
    )
    if not sheets_3ans:
        raise CereqDataMissing(
            f"Pas de feuille 'Données_3ans' dans {xlsx_path.name}. "
            f"Feuilles trouvées : {wb.sheetnames}"
        )

    rows_3ans = _read_cereq_sheet(wb[sheets_3ans])
    rows_6ans = _read_cereq_sheet(wb[sheets_6ans]) if sheets_6ans else {}

    entries: list[dict[str, Any]] = []
    for code, row3 in rows_3ans.items():
        row6 = rows_6ans.get(code, {})
        niveau = CEREQ_CODE_TO_NIVEAU.get(code)
        domaine = CEREQ_DOMAINE_HINTS.get(code)
        entries.append({
            "source": "cereq",
            "cohorte": cohorte,
            "code": code,
            "libelle_menu": row3.pop("Libelle_Menu", None),
            "libelle_complet": row3.pop("Libelle complet", None),
            "niveau": niveau,
            "domaine": domaine,
            "horizon_3ans": _metrics_from_row(row3),
            "horizon_6ans": _metrics_from_row(row6) if row6 else None,
        })
    return entries


def _read_cereq_sheet(ws) -> dict[int, dict[str, Any]]:
    """Lit une feuille Données_* en dict {code: row_as_dict}."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = list(rows[0])
    out: dict[int, dict[str, Any]] = {}
    for raw in rows[1:]:
        record = dict(zip(headers, raw))
        code = record.get("Code")
        if code is None:
            continue
        try:
            code_int = int(code)
        except (TypeError, ValueError):
            continue
        out[code_int] = record
    return out


def _metrics_from_row(row: dict[str, Any]) -> dict[str, Any]:
    """Extrait les métriques numériques d'une row Céreq en ignorant les
    colonnes administratives.

    Les colonnes `eff_pond_*` sont conservées (effectifs pondérés par
    métrique, utiles pour pondération downstream).
    """
    skip = {"Code", "Libelle_Menu", "Libelle complet"}
    return {k: v for k, v in row.items() if k and k not in skip and v is not None}


def collect_cereq_stats(
    raw_dir: Path = RAW_DIR, save: bool = True
) -> list[dict[str, Any]]:
    """Parse les sources Céreq présentes dans `data/raw/`.

    Supporte deux formats :
    1. xlsx OpenData (commité par Matteo) : pattern
       `OpenData_Cereq-Enq_Generation-*.xlsx` — format officiel 2025+
    2. CSV legacy (`data/raw/cereq/*.csv`) : rétro-compatible si Matteo
       télécharge un export custom depuis le portail datavisualisation.

    Si aucune source trouvée → CereqDataMissing avec pointeur TODO.
    """
    # xlsx OpenData (priorité — format officiel)
    xlsxs = sorted(raw_dir.glob("OpenData_Cereq*.xlsx"))
    # Déduplication par taille (les `(1).xlsx` sont des copies inutiles)
    seen_sizes: set[int] = set()
    unique_xlsxs: list[Path] = []
    for p in xlsxs:
        sz = p.stat().st_size
        if sz not in seen_sizes:
            seen_sizes.add(sz)
            unique_xlsxs.append(p)

    all_entries: list[dict[str, Any]] = []
    for p in unique_xlsxs:
        all_entries.extend(parse_cereq_xlsx(p))

    # CSVs legacy (fallback)
    if not all_entries:
        csvs = sorted(RAW_DIR_CSV.glob("*.csv")) if RAW_DIR_CSV.exists() else []
        for p in csvs:
            all_entries.extend(parse_chiffres_cles_csv(p))

    if not all_entries:
        raise CereqDataMissing(
            f"Aucune source Céreq dans {raw_dir} ni {RAW_DIR_CSV}. "
            "Cf docs/TODO_MATTEO_APIS.md §4 pour la procédure."
        )

    if save:
        path = save_processed(all_entries)
        print(f"  [cereq] {len(all_entries)} entrées stats insertion → {path}")
    return all_entries


if __name__ == "__main__":
    try:
        entries = collect_cereq_stats()
        print(f"  [cereq] total entries : {len(entries)}")
    except CereqDataMissing as e:
        print(f"  [cereq] ⚠️  {e}")
        print("  [cereq] Procédure : télécharger manuellement depuis")
        print("  [cereq]   https://www.cereq.fr/datavisualisation/")
        print(f"  [cereq]   vers {RAW_DIR}/")
