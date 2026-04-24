"""INSEE — salaires PCS + taux d'emploi par diplôme (D14 Axe 1, ADR-039 phase c).

Source : https://www.insee.fr (statistique publique, Etalab 2.0).

**Deux datasets scaffoldés ici** :

1. **SALCS (Base Tous Salariés, agrégat PCS × âge × sexe)** — salaires
   nets EQTP médians + déciles par catégorie socioprofessionnelle et
   tranche d'âge. Produit `data/processed/insee_salaires_pcs_age.json`.

2. **Enquête Emploi trimestrielle (taux d'emploi par diplôme × âge)** —
   taux d'emploi, taux de chômage, taux d'activité par niveau de diplôme
   et tranche d'âge. Produit `data/processed/insee_taux_emploi_diplome.json`.

**Particularité ingestion** : INSEE publie ses données via xlsx OpenData
sur insee.fr (pas d'API bulk-download standardisée). Les fichiers sont
livrés par Matteo dans `data/raw/insee/` selon procédure
`docs/TODO_MATTEO_APIS.md` §6.

**Robustesse** : les xlsx INSEE varient en structure (header multi-row,
feuilles "Figure 1..N", cases fusionnées). Les parsers ici acceptent
une `ParserConfig` explicite (sheet_name, header_row, colonnes) pour
s'adapter au format exact du xlsx livré. Exception
`InseeLayoutUnexpected` si la feuille ne contient pas les colonnes
attendues — le message pointe vers les champs requis.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional


RAW_DIR = Path("data/raw/insee")
PROCESSED_SALCS_PATH = Path("data/processed/insee_salaires_pcs_age.json")
PROCESSED_EMPLOI_PATH = Path("data/processed/insee_taux_emploi_diplome.json")


class InseeDataMissing(Exception):
    """XLSX INSEE attendu absent de `data/raw/insee/`. Pointer vers TODO §6."""


class InseeLayoutUnexpected(Exception):
    """Le xlsx a été lu mais son layout ne contient pas les colonnes attendues.

    Le message inclut la liste des headers trouvés et ceux manquants. L'utilisateur
    peut soit :
    - Éditer le xlsx pour renommer les colonnes
    - Fournir une `ParserConfig` personnalisée avec le mapping adapté
    """


# --- Mapping PCS-2003 → domaine OrientIA (high-level + granulaire) ---
#
# La nomenclature INSEE PCS-2003 structure les catégories en 4 niveaux
# (1 / 2 / 3 / 4 chiffres). Pour OrientIA on mappe au niveau 1-2 chiffres
# qui est typiquement présent dans les tableaux SALCS agrégés.

PCS_1CHIFFRE_TO_DOMAINE: dict[str, str] = {
    "1": "agriculture",
    "2": "artisanat_commerce",
    "3": "cadres_prof_sup",
    "4": "prof_intermediaires",
    "5": "employes",
    "6": "ouvriers",
}

PCS_2CHIFFRES_TO_DOMAINE: dict[str, str] = {
    "31": "cadres_prof_lib",       # Professions libérales
    "33": "cadres_fonction_pub",   # Cadres de la fonction publique
    "34": "cadres_enseign_sci",    # Professeurs, professions scientifiques
    "35": "cadres_arts_media",     # Arts, spectacles, information
    "37": "cadres_admin_comm",     # Cadres administratifs et commerciaux d'entreprise
    "38": "cadres_ingenieurs",     # Ingénieurs et cadres techniques d'entreprise
    "42": "prof_inter_enseign",    # Instituteurs et assimilés
    "43": "prof_inter_sante_social", # Professions intermédiaires santé/social
    "44": "prof_inter_cultes",     # Clergé, religieux
    "45": "prof_inter_admin_pub",  # Professions intermédiaires fonction publique
    "46": "prof_inter_admin_ent",  # Professions intermédiaires entreprise
    "47": "prof_inter_technique",  # Techniciens (hors contremaîtres)
    "48": "prof_inter_contremai",  # Contremaîtres, agents de maîtrise
    "52": "employes_fonction_pub",
    "53": "employes_police_militaire",
    "54": "employes_admin_ent",
    "55": "employes_commerce",
    "56": "employes_services_dir", # Personnels des services directs aux particuliers
    "62": "ouvriers_qualifies_indus",
    "63": "ouvriers_qualifies_artisanat",
    "64": "chauffeurs",
    "65": "ouvriers_qualifies_manut_stock_trans",
    "67": "ouvriers_non_qualifies_indus",
    "68": "ouvriers_non_qualifies_artisanat",
    "69": "ouvriers_agricoles",
}


def pcs_to_domaine(pcs: Optional[str]) -> Optional[str]:
    """Renvoie le domaine OrientIA pour un code PCS (2 ou 1 chiffre)."""
    if not pcs:
        return None
    pcs = str(pcs).strip()
    if len(pcs) >= 2 and pcs[:2] in PCS_2CHIFFRES_TO_DOMAINE:
        return PCS_2CHIFFRES_TO_DOMAINE[pcs[:2]]
    if pcs[:1] in PCS_1CHIFFRE_TO_DOMAINE:
        return PCS_1CHIFFRE_TO_DOMAINE[pcs[:1]]
    return None


# --- Mapping niveau diplôme INSEE → niveau OrientIA ---
#
# L'Enquête Emploi INSEE utilise typiquement 6 modalités :
#   - "Sans diplôme ou CEP"
#   - "Brevet des collèges"
#   - "CAP / BEP"
#   - "Baccalauréat général, techno, pro"
#   - "Bac+2"
#   - "Bac+3 ou plus"
#   - "Supérieur long (Bac+5 ou plus)"
#
# Le libellé exact peut varier. Le mapping est conservative : tout label
# non-reconnu → niveau=None, exposé dans le JSON pour investigation.

DIPLOME_TO_NIVEAU: dict[str, str] = {
    # Patterns (lowercase substring match)
    "sans diplôme": "aucun",
    "cep": "aucun",
    "brevet": "brevet",
    "cap": "cap-bep",
    "bep": "cap-bep",
    "bac général": "bac",
    "baccalauréat général": "bac",
    "bac technologique": "bac",
    "bac techno": "bac",
    "bac pro": "bac",
    "baccalauréat pro": "bac",
    "baccalauréat": "bac",  # fallback bac seul
    "bac+2": "bac+2",
    "bac + 2": "bac+2",
    "bac+3": "bac+3",
    "bac + 3": "bac+3",
    "licence": "bac+3",
    "bac+5": "bac+5",
    "bac + 5": "bac+5",
    "master": "bac+5",
    "supérieur long": "bac+5",
    "doctorat": "bac+8",
}


def diplome_to_niveau(label: Optional[str]) -> Optional[str]:
    """Mapping libellé diplôme INSEE → niveau OrientIA."""
    if not label:
        return None
    key = label.lower().strip()
    # Priorité aux patterns les plus spécifiques (par ordre décroissant de longueur)
    for pattern in sorted(DIPLOME_TO_NIVEAU, key=len, reverse=True):
        if pattern in key:
            return DIPLOME_TO_NIVEAU[pattern]
    return None


# --- Parser config (pour adapter aux layouts variables INSEE) ---


@dataclass
class ParserConfig:
    """Configuration d'un parser xlsx INSEE.

    INSEE publie des xlsx avec layouts divers — feuilles "Figure 1/2/3",
    headers multi-row, cases fusionnées. Cette config permet de pointer
    explicitement le parser vers la bonne feuille + ligne d'entête.
    """
    sheet_name: Optional[str] = None  # None = première feuille "Figure" détectée
    header_row: int = 1  # 1-indexed (openpyxl convention)
    skip_rows_after_header: int = 0  # Pour cases fusionnées
    min_cols_required: int = 3


# --- Parser SALCS (salaires PCS × âge × sexe) ---


def parse_salcs_xlsx(
    xlsx_path: Path,
    config: Optional[ParserConfig] = None,
) -> list[dict[str, Any]]:
    """Parse un xlsx INSEE SALCS (Base Tous Salariés agrégat).

    Colonnes attendues (mapping permissif sur nom) :
    - PCS (ou "Catégorie", "pcs2003") : code ou libellé
    - Tranche d'âge (ou "age", "tranche_age") : ex "25-29 ans"
    - Sexe (ou "genre") : "Hommes" / "Femmes" / "Tous" ; optionnel
    - Salaire net EQTP médian (ou "salaire_median", "Q50") : float € annuel
    - Optionnel : 1er décile, 9e décile, effectif

    Layouts INSEE typiques : `Figure 1` ou `Figure 2` selon l'étude.
    Config par défaut : première feuille contenant "Figure" dans le nom,
    header sur ligne 1.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise InseeDataMissing(
            "Package `openpyxl` absent. Installer via `pip install openpyxl`."
        ) from exc

    if not xlsx_path.exists():
        raise InseeDataMissing(
            f"XLSX INSEE SALCS absent : {xlsx_path}. "
            "Cf docs/TODO_MATTEO_APIS.md §6 pour la procédure download."
        )

    cfg = config or ParserConfig()
    wb = load_workbook(xlsx_path, data_only=True)

    sheet_name = cfg.sheet_name
    if not sheet_name:
        # Auto-détect : première feuille "Figure" ou à défaut première feuille
        candidates = [s for s in wb.sheetnames if "figure" in s.lower()]
        sheet_name = candidates[0] if candidates else wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < cfg.header_row:
        raise InseeLayoutUnexpected(
            f"Feuille {sheet_name!r} de {xlsx_path.name} vide ou trop courte."
        )

    header = [str(c or "").strip() for c in rows[cfg.header_row - 1]]
    data_rows = rows[cfg.header_row + cfg.skip_rows_after_header:]

    # Mapping header → champ canonique (case-insensitive substring match)
    col_map = _detect_salcs_columns(header)
    required = {"pcs", "salaire_median"}
    missing = required - set(col_map)
    if missing:
        raise InseeLayoutUnexpected(
            f"Colonnes manquantes dans {xlsx_path.name}/{sheet_name} : {missing}. "
            f"Headers trouvés : {header}. Fournir une ParserConfig custom ou "
            "renommer les colonnes dans le xlsx (cf docstring parse_salcs_xlsx)."
        )

    entries: list[dict[str, Any]] = []
    millesime = _infer_millesime(xlsx_path.name)
    for raw in data_rows:
        if raw is None or all(v is None for v in raw):
            continue
        row = dict(zip(header, raw))
        pcs = row.get(col_map.get("pcs"))
        if pcs is None or str(pcs).strip() == "":
            continue
        entries.append({
            "source": "insee_salcs",
            "millesime": millesime,
            "pcs": str(pcs).strip(),
            "pcs_label": str(row.get(col_map.get("pcs_label"), "") or "").strip() or None,
            "domaine_orientia": pcs_to_domaine(str(pcs)),
            "tranche_age": _clean_str(row.get(col_map.get("tranche_age"))),
            "sexe": _clean_str(row.get(col_map.get("sexe"), "Tous")) or "Tous",
            "salaire_eqtp_net_median_annuel": _safe_int(row.get(col_map.get("salaire_median"))),
            "salaire_eqtp_net_d1": _safe_int(row.get(col_map.get("salaire_d1"))),
            "salaire_eqtp_net_d9": _safe_int(row.get(col_map.get("salaire_d9"))),
            "effectif": _safe_int(row.get(col_map.get("effectif"))),
        })
    return entries


def _detect_salcs_columns(header: list[str]) -> dict[str, str]:
    """Mapping header xlsx → noms canoniques via matching permissif."""
    col_map: dict[str, str] = {}
    for h in header:
        hl = h.lower()
        if not hl:
            continue
        if "pcs" in hl or "catégorie" in hl or "csp" in hl:
            if "libell" in hl or "label" in hl:
                col_map["pcs_label"] = h
            else:
                col_map.setdefault("pcs", h)
        elif "âge" in hl or "age" in hl or "tranche" in hl:
            col_map["tranche_age"] = h
        elif "sexe" in hl or "genre" in hl:
            col_map["sexe"] = h
        elif "médian" in hl or "q50" in hl or "p50" in hl:
            col_map["salaire_median"] = h
        elif "d1" in hl or "p10" in hl or "1er décile" in hl:
            col_map["salaire_d1"] = h
        elif "d9" in hl or "p90" in hl or "9e décile" in hl:
            col_map["salaire_d9"] = h
        elif "effectif" in hl or "nombre" in hl:
            col_map["effectif"] = h
    return col_map


# --- Parser Enquête Emploi (taux emploi par diplôme × âge) ---


def parse_enquete_emploi_xlsx(
    xlsx_path: Path,
    config: Optional[ParserConfig] = None,
) -> list[dict[str, Any]]:
    """Parse un xlsx INSEE Enquête Emploi (taux emploi par diplôme × âge).

    Colonnes attendues :
    - Diplôme (ou "niveau_diplome") : libellé ex "Bac+5 ou plus"
    - Tranche d'âge : ex "15-24", "25-34", "35-49"
    - Taux d'emploi : float 0-100 ou 0-1
    - Optionnel : taux chômage, taux activité, sexe

    Layouts INSEE : souvent 1 feuille par indicateur. Le parser matche le
    taux d'emploi ; les autres taux sont extraits s'ils sont dans la même
    feuille.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise InseeDataMissing(
            "Package `openpyxl` absent. Installer via `pip install openpyxl`."
        ) from exc

    if not xlsx_path.exists():
        raise InseeDataMissing(
            f"XLSX INSEE Enquête Emploi absent : {xlsx_path}. "
            "Cf docs/TODO_MATTEO_APIS.md §6."
        )

    cfg = config or ParserConfig()
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = cfg.sheet_name or wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < cfg.header_row:
        raise InseeLayoutUnexpected(
            f"Feuille {sheet_name!r} de {xlsx_path.name} vide ou trop courte."
        )

    header = [str(c or "").strip() for c in rows[cfg.header_row - 1]]
    data_rows = rows[cfg.header_row + cfg.skip_rows_after_header:]
    col_map = _detect_emploi_columns(header)

    required = {"diplome", "taux_emploi"}
    missing = required - set(col_map)
    if missing:
        raise InseeLayoutUnexpected(
            f"Colonnes manquantes dans {xlsx_path.name}/{sheet_name} : {missing}. "
            f"Headers trouvés : {header}. Fournir ParserConfig custom."
        )

    entries: list[dict[str, Any]] = []
    periode = _infer_periode(xlsx_path.name)
    for raw in data_rows:
        if raw is None or all(v is None for v in raw):
            continue
        row = dict(zip(header, raw))
        diplome = row.get(col_map.get("diplome"))
        if diplome is None or str(diplome).strip() == "":
            continue
        diplome_str = str(diplome).strip()
        entries.append({
            "source": "insee_enquete_emploi",
            "periode": periode,
            "diplome_label": diplome_str,
            "niveau_orientia": diplome_to_niveau(diplome_str),
            "tranche_age": _clean_str(row.get(col_map.get("tranche_age"))),
            "sexe": _clean_str(row.get(col_map.get("sexe"), "Tous")) or "Tous",
            "taux_emploi": _safe_ratio(row.get(col_map.get("taux_emploi"))),
            "taux_chomage": _safe_ratio(row.get(col_map.get("taux_chomage"))),
            "taux_activite": _safe_ratio(row.get(col_map.get("taux_activite"))),
        })
    return entries


def _detect_emploi_columns(header: list[str]) -> dict[str, str]:
    """Mapping header xlsx Enquête Emploi → noms canoniques.

    Ordre des tests critique — "chômage" contient "age", donc le match
    "chomage" doit être tenté AVANT tout match "age"/"âge". De même,
    "emploi" et "activité" sont checkés avant "age" pour éviter les
    collisions sur les colonnes taux.
    """
    col_map: dict[str, str] = {}
    for h in header:
        hl = h.lower()
        if not hl:
            continue
        # Order matters : patterns spécifiques d'abord pour éviter les collisions
        if "chômage" in hl or "chomage" in hl:
            col_map["taux_chomage"] = h
        elif "activité" in hl or "activite" in hl:
            col_map["taux_activite"] = h
        elif "emploi" in hl and "taux" in hl:
            col_map["taux_emploi"] = h
        elif "diplôme" in hl or "diplome" in hl or "niveau" in hl:
            col_map.setdefault("diplome", h)
        elif "sexe" in hl or "genre" in hl:
            col_map["sexe"] = h
        elif "âge" in hl or "tranche" in hl:
            col_map["tranche_age"] = h
    return col_map


# --- Utilitaires génériques ---


def _safe_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        s = str(val).replace("\u202f", "").replace(" ", "").replace("€", "").replace(",", ".").strip()
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _safe_ratio(val: Any) -> Optional[float]:
    """Convertit en ratio [0,1]. Accepte "82,5%" ou "82.5" ou 0.825."""
    if val is None or val == "":
        return None
    try:
        s = str(val).replace("\u202f", "").replace(" ", "").replace("%", "").replace(",", ".").strip()
        f = float(s)
    except (ValueError, TypeError):
        return None
    # Si > 1, c'est un pourcentage (ex 82.5 → 0.825)
    return f / 100.0 if f > 1.5 else f


def _clean_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _infer_millesime(filename: str) -> Optional[str]:
    """Extrait une année (2020-2030) du nom de fichier si présente."""
    import re
    m = re.search(r"(20\d{2})", filename)
    return m.group(1) if m else None


def _infer_periode(filename: str) -> Optional[str]:
    """Extrait une période (YYYY ou YYYYTN) du nom de fichier."""
    import re
    m = re.search(r"(20\d{2}T[1-4])", filename)
    if m:
        return m.group(1)
    return _infer_millesime(filename)


# --- Sauvegarde ---


def save_processed(
    entries: list[dict[str, Any]], path: Path,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return path


# --- Entrée principale ---


def collect_insee_stats(
    raw_dir: Path = RAW_DIR, save: bool = True
) -> dict[str, list[dict[str, Any]]]:
    """Ingère tous les xlsx INSEE présents dans `data/raw/insee/`.

    Pattern attendus :
    - `insee_salaires_pcs*.xlsx` → parse_salcs_xlsx → insee_salaires_pcs_age.json
    - `insee_taux_emploi*.xlsx` ou `insee_enquete_emploi*.xlsx` →
      parse_enquete_emploi_xlsx → insee_taux_emploi_diplome.json

    Si aucun xlsx dispo → raise InseeDataMissing (pas de graceful no-op —
    c'est un scaffold pour quand Matteo livre les fichiers, si la fonction
    s'execute c'est qu'on veut les données).

    Retourne un dict {"salcs": [...], "enquete_emploi": [...]}.
    """
    if not raw_dir.exists():
        raise InseeDataMissing(
            f"Dossier {raw_dir} absent. Créer avec `mkdir -p {raw_dir}` et y "
            "déposer les xlsx INSEE (cf docs/TODO_MATTEO_APIS.md §6)."
        )

    salcs_paths = sorted(raw_dir.glob("insee_salaires_pcs*.xlsx"))
    emploi_paths = sorted(raw_dir.glob("insee_taux_emploi*.xlsx")) + sorted(raw_dir.glob("insee_enquete_emploi*.xlsx"))

    results: dict[str, list[dict[str, Any]]] = {"salcs": [], "enquete_emploi": []}

    for p in salcs_paths:
        results["salcs"].extend(parse_salcs_xlsx(p))
    for p in emploi_paths:
        results["enquete_emploi"].extend(parse_enquete_emploi_xlsx(p))

    if not results["salcs"] and not results["enquete_emploi"]:
        raise InseeDataMissing(
            f"Aucun xlsx INSEE trouvé dans {raw_dir}. "
            "Attendu : insee_salaires_pcs*.xlsx + insee_taux_emploi*.xlsx. "
            "Cf docs/TODO_MATTEO_APIS.md §6."
        )

    if save:
        if results["salcs"]:
            save_processed(results["salcs"], PROCESSED_SALCS_PATH)
            print(f"  [insee] SALCS : {len(results['salcs'])} entrées → {PROCESSED_SALCS_PATH}")
        if results["enquete_emploi"]:
            save_processed(results["enquete_emploi"], PROCESSED_EMPLOI_PATH)
            print(f"  [insee] Enquête Emploi : {len(results['enquete_emploi'])} entrées → {PROCESSED_EMPLOI_PATH}")

    return results


if __name__ == "__main__":
    try:
        stats = collect_insee_stats()
        print(f"  [insee] total : {len(stats['salcs'])} SALCS + {len(stats['enquete_emploi'])} Enquête Emploi")
    except InseeDataMissing as e:
        print(f"  [insee] ⚠️  {e}")
        print("  [insee] Procédure Matteo : cf docs/TODO_MATTEO_APIS.md §6")
