"""DARES "Les métiers en 2030" — projections recrutement par métier × région.

Source : https://dares.travail-emploi.gouv.fr/publication/les-metiers-en-2030-quels-desequilibres-potentiels
Licence : Etalab 2.0 (statistique publique).
Format : tableaux Excel annexes publiés avec le rapport DARES 2022.

**Dimension unique apportée à OrientIA** (ADR-039) :

- **Prospective** : les autres sources OrientIA (Céreq, InserSup,
  Inserjeunes, France Travail) sont **rétrospectives** (insertion passée).
  DARES Métiers 2030 donne des **projections de recrutement** jusqu'en
  2030 par métier (FAP 87 postes × ROME) × région × scénario.
- **Critique pour l'orientation** : un·e lycéen·ne qui choisit une
  formation en 2025 sort en 2028-2030. La question "quels métiers
  vont recruter quand je sortirai" n'est adressable qu'avec cette
  source. Les autres répondent "quels métiers recrutaient il y a
  5 ans" — utile mais pas suffisant.

**Statut** : 📋 **MANUEL** — parser scaffold prêt, attend l'upload par
Matteo des XLSX d'annexes du rapport DARES (procédure
`docs/TODO_MATTEO_APIS.md` §8 à créer).

Les XLSX attendus sont typiquement :
- `Metiers_2030_FAP87_Projections.xlsx` : projections par famille
  professionnelle (FAP 87 postes — la nomenclature métiers DARES)
- `Metiers_2030_Region.xlsx` : projections par région × grand secteur
- `Metiers_2030_FAP_Region.xlsx` (optionnel) : croisement FAP × région

**Robustesse** : xlsx DARES ont souvent des layouts complexes (header
multi-row, cellules fusionnées pour les scénarios). Config explicite
via `ParserConfig` dataclass (comme pour INSEE).
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional


RAW_DIR = Path("data/raw/dares")
PROCESSED_PATH = Path("data/processed/dares_metiers_2030.json")


class DaresDataMissing(Exception):
    """XLSX DARES Métiers 2030 attendu absent de `data/raw/dares/`.

    Pointer vers `docs/TODO_MATTEO_APIS.md` §8 pour procédure upload.
    """


class DaresLayoutUnexpected(Exception):
    """Le xlsx a été lu mais son layout ne contient pas les colonnes attendues."""


# --- Mapping FAP (Famille Professionnelle) DARES → domaine OrientIA ---
#
# La nomenclature FAP 87 groupe les métiers en 22 grandes familles
# (A. Agriculture, B. BTP, C. Ingénieurs industrie, D. Banque/assurance,
# E. Informatique/télécoms, F. Électricité/électronique, G. Mécanique,
# H. Industrie process, J. Maintenance, K. Transports/logistique,
# L. Artisanat, M. Gestion/administration, N. Commerce, P. Hôtellerie,
# Q. Politique/culture, R. Santé social, S. Enseignement, T. Service
# personnes, U. Communication/culture, V. Santé médicale, W. Armée,
# X. Non classable).
FAP_PREFIX_TO_DOMAINE: dict[str, str] = {
    "A": "agriculture",
    "B": "btp",
    "C": "ingenierie_industrielle",
    "D": "banque_assurance",
    "E": "data_ia",  # Info + télécoms
    "F": "ingenierie_industrielle",
    "G": "ingenierie_industrielle",
    "H": "industrie_process",
    "J": "maintenance",
    "K": "logistique_transport",
    "L": "artisanat",
    "M": "eco_gestion",
    "N": "commerce",
    "P": "hotellerie_restauration",
    "Q": "services_publics",
    "R": "social",
    "S": "enseignement",
    "T": "service_personnes",
    "U": "communication",
    "V": "sante",
    "W": "armee_securite",
}


def fap_to_domaine(fap_code: Optional[str]) -> Optional[str]:
    """Renvoie le domaine OrientIA pour un code FAP (1 lettre + 2 chiffres)."""
    if not fap_code:
        return None
    code = str(fap_code).strip().upper()
    return FAP_PREFIX_TO_DOMAINE.get(code[:1]) if code else None


# --- Parser config ---


@dataclass
class ParserConfig:
    """Configuration parser xlsx DARES."""
    sheet_name: Optional[str] = None  # None = première feuille avec "FAP" ou "projection" dans le nom
    header_row: int = 1  # 1-indexed
    skip_rows_after_header: int = 0


# --- Parser XLSX projections FAP ---


def parse_dares_fap_xlsx(
    xlsx_path: Path,
    config: Optional[ParserConfig] = None,
) -> list[dict[str, Any]]:
    """Parse un xlsx DARES Métiers 2030 projections par FAP.

    Colonnes attendues (matching permissif sur nom) :
    - Code FAP (ou "FAP", "code_fap") : ex "E1Z80" ou "E1"
    - Libellé FAP : nom de la famille pro
    - Postes à pourvoir (créations / départs / total) : par scénario
      DARES (central / favorable / défavorable)
    - Horizon : 2020-2030 typiquement

    Layout variable : DARES publie plusieurs variantes de scénario dans
    la même feuille avec des cellules fusionnées. La config permet de
    pointer vers la bonne feuille et ligne d'entête.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DaresDataMissing(
            "Package `openpyxl` absent. Installer via `pip install openpyxl`."
        ) from exc

    if not xlsx_path.exists():
        raise DaresDataMissing(
            f"XLSX DARES absent : {xlsx_path}. Cf docs/TODO_MATTEO_APIS.md §8."
        )

    cfg = config or ParserConfig()
    wb = load_workbook(xlsx_path, data_only=True)

    sheet_name = cfg.sheet_name
    if not sheet_name:
        # Auto-détect : première feuille avec "FAP" ou "projection" dans le nom
        candidates = [s for s in wb.sheetnames if "fap" in s.lower() or "projection" in s.lower()]
        sheet_name = candidates[0] if candidates else wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < cfg.header_row:
        raise DaresLayoutUnexpected(
            f"Feuille {sheet_name!r} de {xlsx_path.name} vide ou trop courte."
        )

    header = [str(c or "").strip() for c in rows[cfg.header_row - 1]]
    data_rows = rows[cfg.header_row + cfg.skip_rows_after_header:]
    col_map = _detect_fap_columns(header)

    required = {"fap_code", "libelle"}
    missing = required - set(col_map)
    if missing:
        raise DaresLayoutUnexpected(
            f"Colonnes manquantes dans {xlsx_path.name}/{sheet_name} : {missing}. "
            f"Headers trouvés : {header}. Fournir ParserConfig custom."
        )

    entries: list[dict[str, Any]] = []
    for raw in data_rows:
        if raw is None or all(v is None for v in raw):
            continue
        row = dict(zip(header, raw))
        fap_code = row.get(col_map.get("fap_code"))
        if fap_code is None or str(fap_code).strip() == "":
            continue
        entries.append({
            "source": "dares_metiers_2030",
            "horizon": "2020-2030",
            "fap_code": str(fap_code).strip(),
            "fap_libelle": _clean_str(row.get(col_map.get("libelle"))),
            "domaine_orientia": fap_to_domaine(str(fap_code)),
            "postes_a_pourvoir_total": _safe_int(row.get(col_map.get("postes_total"))),
            "postes_creation": _safe_int(row.get(col_map.get("postes_creation"))),
            "postes_depart_retraite": _safe_int(row.get(col_map.get("postes_depart"))),
            "scenario": _clean_str(row.get(col_map.get("scenario"), "central")) or "central",
        })
    return entries


def _detect_fap_columns(header: list[str]) -> dict[str, str]:
    """Mapping header xlsx → noms canoniques (matching permissif)."""
    col_map: dict[str, str] = {}
    for h in header:
        hl = h.lower()
        if not hl:
            continue
        if "fap" in hl and ("code" in hl or len(hl) < 10):
            col_map.setdefault("fap_code", h)
        elif "libell" in hl or "intitul" in hl or "nom" in hl:
            col_map.setdefault("libelle", h)
        elif "création" in hl or "creation" in hl or "créé" in hl:
            col_map["postes_creation"] = h
        elif "départ" in hl or "depart" in hl or "retraite" in hl:
            col_map["postes_depart"] = h
        elif "total" in hl or "pourvoir" in hl:
            col_map["postes_total"] = h
        elif "scénario" in hl or "scenario" in hl:
            col_map["scenario"] = h
    return col_map


# --- Parser XLSX projections régionales ---


def parse_dares_region_xlsx(
    xlsx_path: Path,
    config: Optional[ParserConfig] = None,
) -> list[dict[str, Any]]:
    """Parse un xlsx DARES Métiers 2030 projections par RÉGION.

    Colonnes attendues :
    - Région (ou "code_region") : libellé région française
    - Postes à pourvoir 2020-2030 total / création / départ
    - Optionnel : scénario, secteur d'activité

    Format typique : 13 régions métropolitaines + DOM + France entière.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DaresDataMissing(
            "Package `openpyxl` absent. Installer via `pip install openpyxl`."
        ) from exc

    if not xlsx_path.exists():
        raise DaresDataMissing(
            f"XLSX DARES région absent : {xlsx_path}. Cf docs/TODO_MATTEO_APIS.md §8."
        )

    cfg = config or ParserConfig()
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_name = cfg.sheet_name or wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise DaresLayoutUnexpected(f"Feuille {sheet_name} vide")

    header = [str(c or "").strip() for c in rows[cfg.header_row - 1]]
    data_rows = rows[cfg.header_row + cfg.skip_rows_after_header:]
    col_map = _detect_region_columns(header)

    required = {"region"}
    missing = required - set(col_map)
    if missing:
        raise DaresLayoutUnexpected(
            f"Colonnes manquantes : {missing}. Headers : {header}."
        )

    entries: list[dict[str, Any]] = []
    for raw in data_rows:
        if raw is None or all(v is None for v in raw):
            continue
        row = dict(zip(header, raw))
        region = row.get(col_map.get("region"))
        if region is None or str(region).strip() == "":
            continue
        entries.append({
            "source": "dares_metiers_2030_region",
            "horizon": "2020-2030",
            "region": str(region).strip(),
            "secteur": _clean_str(row.get(col_map.get("secteur"))),
            "postes_a_pourvoir_total": _safe_int(row.get(col_map.get("postes_total"))),
            "postes_creation": _safe_int(row.get(col_map.get("postes_creation"))),
            "postes_depart_retraite": _safe_int(row.get(col_map.get("postes_depart"))),
            "scenario": _clean_str(row.get(col_map.get("scenario"), "central")) or "central",
        })
    return entries


def _detect_region_columns(header: list[str]) -> dict[str, str]:
    col_map: dict[str, str] = {}
    for h in header:
        hl = h.lower()
        if not hl:
            continue
        if "région" in hl or "region" in hl:
            col_map.setdefault("region", h)
        elif "secteur" in hl:
            col_map["secteur"] = h
        elif "création" in hl or "creation" in hl:
            col_map["postes_creation"] = h
        elif "départ" in hl or "retraite" in hl:
            col_map["postes_depart"] = h
        elif "total" in hl or "pourvoir" in hl:
            col_map["postes_total"] = h
        elif "scénario" in hl or "scenario" in hl:
            col_map["scenario"] = h
    return col_map


# --- Utilitaires ---


def _safe_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val).replace("\u202f", "").replace(" ", "").replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _clean_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def save_processed(
    entries: list[dict[str, Any]], path: Path = PROCESSED_PATH,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return path


def collect_dares_metiers_2030(
    raw_dir: Path = RAW_DIR, save: bool = True
) -> list[dict[str, Any]]:
    """Parse tous les xlsx DARES Métiers 2030 présents dans `data/raw/dares/`.

    Pattern attendus :
    - `*FAP*.xlsx` ou `*Projections*.xlsx` → parse_dares_fap_xlsx
    - `*Region*.xlsx` → parse_dares_region_xlsx

    Si aucun xlsx → DaresDataMissing.
    """
    if not raw_dir.exists():
        raise DaresDataMissing(
            f"Dossier {raw_dir} absent. Créer avec `mkdir -p {raw_dir}` puis "
            "y déposer les xlsx DARES (cf docs/TODO_MATTEO_APIS.md §8)."
        )

    entries: list[dict[str, Any]] = []

    # FAP projections — set() pour dédupliquer si un fichier match plusieurs
    # patterns (ex "Metiers2030_FAP_Projections.xlsx" match *FAP* ET
    # *Projections*)
    fap_candidates = set(raw_dir.glob("*FAP*.xlsx")) | set(raw_dir.glob("*Projections*.xlsx"))
    for p in sorted(fap_candidates):
        if "region" in p.name.lower():
            continue  # éviter double-comptage si nom ambigu
        entries.extend(parse_dares_fap_xlsx(p))

    # Régions
    for p in sorted(raw_dir.glob("*Region*.xlsx")):
        entries.extend(parse_dares_region_xlsx(p))

    if not entries:
        raise DaresDataMissing(
            f"Aucun xlsx DARES Métiers 2030 trouvé dans {raw_dir}. "
            "Attendu : *FAP*.xlsx + *Region*.xlsx. "
            "Cf docs/TODO_MATTEO_APIS.md §8."
        )

    if save:
        p = save_processed(entries)
        print(f"  [dares] {len(entries)} projections sauvées → {p}")

    return entries


if __name__ == "__main__":
    try:
        entries = collect_dares_metiers_2030()
        print(f"  [dares] total : {len(entries)} projections")
    except DaresDataMissing as e:
        print(f"  [dares] ⚠️  {e}")
        print("  [dares] Procédure : cf docs/TODO_MATTEO_APIS.md §8")
